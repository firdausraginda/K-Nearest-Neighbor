from openpyxl import load_workbook
import math
import xlwt

# Load in the workbook
wb = load_workbook('./dataset.xlsx')

# Initialize a workbook
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook
sheet1 = book.add_sheet("Jawaban tupro 3 Raginda")

# Get a sheet by name
sheet = wb.get_sheet_by_name('DataTrain')
sheet2 = wb.get_sheet_by_name('DataTest')

listLike = []
listProv = []
listKomen = []
listEmo = []
eDistance = []
eHoax = []
e = []
k = []
h = []

for i in range(2, 1002):
    for j in range(2, 4002):
        hasilLike = ((sheet2.cell(row=i, column=2).value) - (sheet.cell(row=j, column=2).value))
        # print("hasil pengurangan like : ",hasilLike)
        listLike.append(hasilLike * hasilLike)
        # print("hasil kuadrat like: ",arrayLike[j-2])

        hasilProv = ((sheet2.cell(row=i, column=3).value) - (sheet.cell(row=j, column=3).value))
        # print("hasil pengurangan prov : ",hasilProv)
        listProv.append(hasilProv * hasilProv)
        # print("hasil kuadrat prov: ",arrayProv[j-2])

        hasilKomen = ((sheet2.cell(row=i, column=4).value) - (sheet.cell(row=j, column=4).value))
        # print("hasil pengurangan like : ",hasilLike)
        listKomen.append(hasilKomen * hasilKomen)
        # print("hasil kuadrat komen: ",arrayKomen[j-2])

        hasilEmo = ((sheet2.cell(row=i, column=5).value) - (sheet.cell(row=j, column=5).value))
        # print("hasil pengurangan like : ",hasilLike)
        listEmo.append(hasilEmo * hasilEmo)
        # print("hasil kuadrat emosi: ",arrayEmo[j-2])

        jumlahE = listLike[-1] + listProv[-1] + listKomen[-1] + listEmo[-1]
        eDistance.append(math.sqrt(jumlahE))
        eHoax.append(sheet.cell(row=j, column=6).value)
        e.append([eDistance[-1], eHoax[-1]])

        if j==4001:
            # print("jumlah e: ", len(e))
            e.sort()
            k = e[0:101]

            noHoax = 0
            yesHoax = 0

            for x in range(len(k)):
                if k[x][1] == 0:
                    noHoax = noHoax + 1
                else:
                    yesHoax = yesHoax + 1

            hoax=0
            if noHoax > yesHoax:
                hoax = 0
            else:
                hoax = 1

            h.append(hoax)

            e.clear()
            k.clear()

# for p in range(len(h)):
    # print("distance", e[i])
    # print("yeye", k)
    # print("lala", hoax[i])

for p in range(len(h)):
    print("nilai hoax ke-",p+1,": ", h[p])
    # print(h[p])
    # sheet1.write(i, 0, h[p])

# Write to the sheet of the workbook
# for i in range(len(h)):
    # sheet1.write(i, 0, h[i])

# Save the workbook
# book.save("jawabanDataTestxxx.xls")